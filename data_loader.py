#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import annotations

import logging
import os
import unicodedata
from typing import Iterable, Callable, Optional

import pandas as pd

logger = logging.getLogger(__name__)


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    # normaliza espacios por guiones bajos y quita espacios
    df = df.copy()
    df.columns = [c.strip().replace(" ", "_") for c in df.columns]

    def _canon(s: str) -> str:
        s2 = unicodedata.normalize('NFD', s)
        s2 = ''.join(ch for ch in s2 if unicodedata.category(ch) != 'Mn')
        return s2.lower()

    cols = list(df.columns)
    canon_to_actual = {_canon(c): c for c in cols}

    def _map_to(target: str, candidates: Iterable[str]) -> None:
        if target in df.columns:
            return
        for k in candidates:
            if k in canon_to_actual:
                df.rename(columns={canon_to_actual[k]: target}, inplace=True)
                return

    # Producto / Nombre
    _map_to("Nombre_del_Producto", ["producto", "nombre_del_producto"])
    # Fecha vencimiento
    _map_to("Fecha_de_Vencimiento", ["fecha_de_vencimiento", "fecha_vencimiento"])
    # Cantidad/Stock disponible
    _map_to("Cantidad_Disponible", ["saldo_stock", "cantidad_disponible"])  # preferimos Cantidad_Disponible
    # Ubicacion (con o sin acento)
    if "Ubicacion" not in df.columns:
        for cand in ["ubicacion", "ubicación"]:
            if cand in canon_to_actual:
                df.rename(columns={canon_to_actual[cand]: "Ubicacion"}, inplace=True)
                break
    # Familia y Subfamilia
    _map_to("Familia", ["familia"])
    _map_to("Subfamilia", ["subfamilia"])
    # Código
    _map_to("Codigo", ["codigo", "código"])
    # Unidad
    _map_to("Unidad", ["unidad"])
    # Unidad de negocio
    _map_to("Unidad_de_negocio", ["unidad_de_negocio"]) 
    # Bodega
    _map_to("Bodega", ["bodega"]) 
    # N° Serie y variantes
    for cand in ["n°_serie", "nº_serie", "n_serie", "numero_de_serie", "nro_serie"]:
        if "N_Serie" not in df.columns and cand in canon_to_actual:
            df.rename(columns={canon_to_actual[cand]: "N_Serie"}, inplace=True)
            break
    # Por llegar y Reserva
    _map_to("Por_llegar", ["por_llegar"]) 
    _map_to("Reserva", ["reserva"]) 

    return df


def _filter_by_area(df: pd.DataFrame, area_name: str) -> pd.DataFrame:
    # detecta posibles nombres de columna de área
    for candidate in ("Área", "�?rea", "Area"):
        if candidate in df.columns:
            return df[df[candidate] == area_name].copy()
    # si no existe columna, devolver todo
    return df.copy()


def _read_excel_stream(
    file_path: str,
    progress_cb: Optional[Callable[[int, Optional[int], str], None]] = None,
    chunk_size: int = 2000,
) -> pd.DataFrame:
    try:
        import openpyxl  # type: ignore
    except Exception as exc:
        raise RuntimeError("openpyxl no disponible para lectura por slots") from exc

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        total_rows = max((ws.max_row or 1) - 1, 0)
        if progress_cb:
            progress_cb(0, total_rows, "Leyendo archivo...")
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return pd.DataFrame()
        headers = [str(h).strip() if h is not None else "" for h in headers]
        frames = []
        batch = []
        processed = 0
        for row in rows_iter:
            row_dict = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            batch.append(row_dict)
            processed += 1
            if len(batch) >= chunk_size:
                frames.append(pd.DataFrame(batch))
                batch = []
                if progress_cb:
                    progress_cb(processed, total_rows, "Leyendo archivo...")
        if batch:
            frames.append(pd.DataFrame(batch))
        if progress_cb:
            progress_cb(total_rows, total_rows, "Procesando datos...")
        if not frames:
            return pd.DataFrame(columns=headers)
        return pd.concat(frames, ignore_index=True)
    finally:
        try:
            wb.close()
        except Exception:
            pass


def load_inventory(
    file_path: str,
    area_filter: str | None = None,
    progress_cb: Optional[Callable[[int, Optional[int], str], None]] = None,
    chunk_size: int = 2000,
) -> pd.DataFrame:
    if not os.path.exists(file_path):
        raise FileNotFoundError(file_path)

    logger.info("Cargando inventario desde %s", file_path)
    ext = os.path.splitext(file_path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xltx", ".xltm"):
        try:
            df = _read_excel_stream(file_path, progress_cb=progress_cb, chunk_size=chunk_size)
        except Exception:
            if progress_cb:
                progress_cb(0, None, "Leyendo archivo...")
            df = pd.read_excel(file_path)
    else:
        if progress_cb:
            progress_cb(0, None, "Leyendo archivo...")
        df = pd.read_excel(file_path)
    df = _normalize_columns(df)

    required = ["Nombre_del_Producto", "Lote", "Fecha_de_Vencimiento", "Cantidad_Disponible"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise KeyError(
            "Faltan columnas requeridas: " + ", ".join(m.replace("_", " ") for m in missing)
        )

    if area_filter:
        df = _filter_by_area(df, area_filter)

    # Campos normalizados
    df = df.copy()
    df["Stock"] = df["Cantidad_Disponible"].fillna(0).astype(int)
    df["Vencimiento"] = pd.to_datetime(
        df["Fecha_de_Vencimiento"], errors="coerce", dayfirst=True
    ).dt.strftime("%Y-%m-%d")

    # Ubicación opcional
    if "Ubicacion" not in df.columns:
        df["Ubicacion"] = ""

    # Columnas para la UI (incluye familia/subfamilia para agrupar)
    desired = [
        "Familia",
        "Subfamilia",
        "Codigo",
        "Nombre_del_Producto",
        "Unidad",
        "Unidad_de_negocio",
        "Bodega",
        "Ubicacion",
        "N_Serie",
        "Lote",
        "Vencimiento",
        "Por_llegar",
        "Reserva",
        "Stock",
    ]

    for c in desired:
        if c not in df.columns:
            if c in ("Por_llegar", "Reserva", "Stock"):
                df[c] = 0
            else:
                df[c] = ""

    logger.info(
        "Inventario cargado: %d filas (filtro área=%s)",
        len(df),
        area_filter or "N/A",
    )
    return df[desired]
